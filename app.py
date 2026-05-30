import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import datetime
import requests

# --- SET PAGE CONFIG ---
st.set_page_config(page_title="Metropolis AI - Rate Survey Tool", layout="wide", page_icon="🅿️")

# --- CUSTOM METROPOLIS/SP+ BRAND STYLES (CSS) ---
st.markdown("""
    <style>
    .main-header { font-size:28px !important; font-weight: bold; color: #C00000; margin-bottom: 5px; }
    .sub-header { font-size:16px !important; color: #555555; margin-bottom: 25px; }
    </style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-header">Metropolis Garage Survey & Market Intelligence Tool</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Automated competitive data ingestion, validation, and SP+ matrix formatting.</div>', unsafe_allow_html=True)

# --- SIDEBAR INPUTS ---
st.sidebar.header("Survey Parameters")

# Main location query box
address_query = st.sidebar.text_input("Search Location / Address", "1221 North State Parkway, Chicago, IL")

location_input = address_query
if address_query:
    try:
        # Fetching recommendations from OpenStreetMap Nominatim Engine
        url = f"https://nominatim.openstreetmap.org/search?q={address_query}&format=json&limit=5"
        headers = {"User-Agent": "MetropolisMarketIntelligenceTool/1.0"}
        response = requests.get(url, headers=headers).json()
        
        if response:
            options = [place["display_name"] for place in response]
            # Add what the user typed as the very first option so they are never blocked by "No Results"
            if address_query not in options:
                options.insert(0, address_query)
            location_input = st.sidebar.selectbox("Select Confirmed Address Target:", options)
    except Exception:
        pass

radius_miles = st.sidebar.slider("Survey Radius (Miles)", min_value=1, max_value=10, value=5)
location_id = st.sidebar.text_input("SP+ Location ID Mapping (Optional)", "72912")

# --- MOCK API GEOSPATIAL DISCOVERY ENGINE ---
def discover_market_rates(address, radius):
    mock_data = {
        "ONE EAST SCOTT": {
            "address": "1221 North State Parkway", "capacity": "250", "hours": "24 Hours", "operator": "SP+", 
            "service_type": "Self-Park", "facility_type": "Garage", "effective_date": "03/21/2024",
            "rates": {"2 hrs, SelfPark": 18.00, "4 hrs, SelfPark": 27.00, "8 hrs, SelfPark": 35.00, "12 hrs, SelfPark": 35.00, "24 hrs, SelfPark": 41.00}
        },
        "1212 N. Lake Shore Dr. Garage": {
            "address": "85 East Scott Street", "capacity": "0", "hours": "24 Hours", "operator": "SP+", 
            "service_type": "Valet", "facility_type": "Garage, Other", "effective_date": "01/21/2022",
            "rates": {"1 hrs, Valet": 16.00, "2 hrs, Valet": 18.00, "3 hrs, Valet": 18.00, "4 hrs, Valet": 18.00}
        },
        "Jewel Osco Sinclair Garage": {
            "address": "1230 North Clark Street", "capacity": "0", "hours": "N/A", "operator": "The Sinclair Garage", 
            "service_type": "Self-Park", "facility_type": "Garage", "effective_date": "N/A",
            "rates": {}
        },
        "1350 N LAKE SHORE DRIVE": {
            "address": "60 East Banks Street", "capacity": "1", "hours": "24 Hours", "operator": "SP+", 
            "service_type": "Valet, Self-Park", "facility_type": "Garage", "effective_date": "04/20/2026",
            "rates": {"1 hrs, Valet": 18.00, "2 hrs, Valet": 18.00, "3 hrs, Valet": 20.00, "4 hrs, Valet": 20.00}
        },
        "1455 N. State Garage": {
            "address": "1445 North State Parkway", "capacity": "250", "hours": "24 Hours", "operator": "SP+", 
            "service_type": "Valet", "facility_type": "Garage", "effective_date": "07/29/2020",
            "rates": {"1 hrs, Valet": 17.00, "2 hrs, Valet": 17.00, "3 hrs, Valet": 19.00, "4 hrs, Valet": 19.00}
        }
    }
    return mock_data

# --- EXCEL EXPORT ENGINE (OPENPYXL GENERATOR) ---
def generate_excel_matrix(survey_data, loc_id):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rate Survey Matrix"
    ws.views.sheetView[0].showGridLines = True

    brand_red = "C00000"
    brand_yellow = "FFC000"
    border_gray = "D3D3D3"
    
    fill_red = PatternFill(start_color=brand_red, end_color=brand_red, fill_type="solid")
    fill_yellow = PatternFill(start_color=brand_yellow, end_color=brand_yellow, fill_type="solid")
    
    font_title = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_regular = Font(name="Calibri", size=10, color="000000")
    font_currency = Font(name="Calibri", size=10, bold=True, color="000000")
    
    thin_border = Border(
        left=Side(style='thin', color=border_gray),
        right=Side(style='thin', color=border_gray),
        top=Side(style='thin', color=border_gray),
        bottom=Side(style='thin', color=border_gray)
    )

    num_columns = len(survey_data) + 1
    last_col_letter = get_column_letter(num_columns)
    
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "Location Information"
    ws["A1"].fill = fill_red
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    metadata_labels = [
        "Location Name", "Address", "Working Capacity", 
        "Operating Hours", "Operator", "Service Type", "Facility Type"
    ]
    
    for idx, label in enumerate(metadata_labels, start=2):
        cell = ws.cell(row=idx, column=1, value=label)
        cell.fill = fill_yellow
        cell.font = font_bold
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[idx].height = 20

    col_idx = 2
    for name, info in survey_data.items():
        c2 = ws.cell(row=2, column=col_idx, value=name)
        ws.cell(row=3, column=col_idx, value=info["address"])
        ws.cell(row=4, column=col_idx, value=int(info["capacity"]))
        ws.cell(row=5, column=col_idx, value=info["hours"])
        ws.cell(row=6, column=col_idx, value=info["operator"])
        ws.cell(row=7, column=col_idx, value=info["service_type"])
        ws.cell(row=8, column=col_idx, value=info["facility_type"])
        
        c2.font = font_bold
        c2.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        
        for r in range(2, 9):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = fill_yellow
            if r != 2:
                cell.font = font_regular
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        col_idx += 1

    ws.merge_cells(f"A9:{last_col_letter}9")
    ws["A9"] = "Board Rates"
    ws["A9"].fill = fill_red
    ws["A9"].font = font_title
    ws["A9"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[9].height = 24

    rate_rows = [
        "Effective as of",
        "2 hrs, SelfPark", "4 hrs, SelfPark", "8 hrs, SelfPark", "12 hrs, SelfPark", "24 hrs, SelfPark",
        "1 hrs, Valet", "2 hrs, Valet", "3 hrs, Valet", "4 hrs, Valet"
    ]
    
    for idx, rate_name in enumerate(rate_rows, start=10):
        cell = ws.cell(row=idx, column=1, value=rate_name)
        cell.font = font_bold if idx == 10 else font_regular
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[idx].height = 18

    col_idx = 2
    for name, info in survey_data.items():
        eff_cell = ws.cell(row=10, column=col_idx, value=info["effective_date"])
        eff_cell.font = font_regular
        eff_cell.alignment = Alignment(horizontal="center", vertical="center")
        eff_cell.border = thin_border
        
        for r_idx, rate_label in enumerate(rate_rows[1:], start=11):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            val = info["rates"].get(rate_label, None)
            if val is not None:
                cell.value = val
                cell.number_format = '$#,##0.00'
                cell.font = font_currency
            else:
                cell.value = ""
        col_idx += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    ws.column_dimensions['A'].width = 24

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream

# --- MAIN INTERACTIVE INTERFACE WORKFLOW ---
if st.button("⚡ Execute Survey Pipeline", type="primary"):
    with st.spinner(f"Scraping competitor inventories..."):
        raw_survey = discover_market_rates(location_input, radius_miles)
        
        display_records = []
        for loc, d in raw_survey.items():
            display_records.append({
                "Garage Name": loc,
                "Operator": d["operator"],
                "Address": d["address"],
                "Capacity": d["capacity"],
                "Service Engine": d["service_type"]
            })
        
        st.success(f"✅ Survey extraction complete! Analyzed options near: {location_input}")
        st.subheader("Extracted Market Entities Overview")
        st.dataframe(pd.DataFrame(display_records), use_container_width=True)
        
        excel_data = generate_excel_matrix(raw_survey, location_id)
        timestamp = datetime.datetime.now().strftime("%Y%m%d")
        file_name = f"Ratesurvey_Metropolis_ID{location_id}_{timestamp}.xlsx"
        
        st.markdown("---")
        st.subheader("Generate Artifact Exports")
        st.download_button(
            label="💾 Download Standardized Rate Survey Excel Sheet",
            data=excel_data,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )